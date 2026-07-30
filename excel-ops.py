"""
Universal QA Toolkit — Excel ETL Operations Engine (excel-ops.py)
Supports: Reading, Writing, Merging, Deduplicating, Cleaning, and Analyzing Excel/CSV workbooks.
Author: Azeez
"""

import os
import sys
import json
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def read_excel(file_path, sheet_name=0, rows=None, fmt="json", stats=False):
    """Read Excel or CSV file and return formatted data or print statistics."""
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.", file=sys.stderr)
        sys.exit(1)

    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path, dtype=str)
    else:
        df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)

    df = df.fillna("")

    if stats:
        info = {
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "columns": list(df.columns),
            "null_counts": {col: int((df[col] == "").sum()) for col in df.columns},
            "distinct_counts": {col: int(df[col].nunique()) for col in df.columns}
        }
        print(json.dumps(info, indent=2))
        return

    if rows:
        try:
            if "-" in str(rows):
                start, end = map(int, str(rows).split("-"))
                df = df.iloc[start-1:end]
            else:
                df = df.iloc[:int(rows)]
        except Exception as e:
            print(f"Warning: Invalid row slice '{rows}', returning all rows: {e}", file=sys.stderr)

    records = df.to_dict(orient="records")
    if fmt == "json":
        print(json.dumps(records, indent=2, ensure_ascii=False))
    elif fmt == "csv":
        df.to_csv(sys.stdout, index=False)

def write_excel(input_json, output_file, sheet_name="Data"):
    """Write JSON records to a professionally styled Excel file."""
    if os.path.exists(input_json):
        with open(input_json, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        try:
            data = json.loads(input_json)
        except Exception:
            print(f"Error: Cannot parse JSON input from '{input_json}'.", file=sys.stderr)
            sys.exit(1)

    if not isinstance(data, list) or not data:
        print("Error: Input data must be a non-empty list of dictionaries.", file=sys.stderr)
        sys.exit(1)

    df = pd.DataFrame(data).fillna("")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(df.columns)
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_data in df.values.tolist():
        ws.append([str(v) for v in row_data])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(os.path.abspath(output_file)), exist_ok=True)
    wb.save(output_file)
    print(f"Successfully wrote {len(df)} rows to '{output_file}' (Sheet: {sheet_name})")

def merge_excel(files, dedup_key=None, output_file="merged_output.xlsx"):
    """Merge multiple Excel/CSV files and optionally deduplicate."""
    dfs = []
    for f in files:
        if not os.path.exists(f):
            print(f"Warning: File '{f}' does not exist, skipping.", file=sys.stderr)
            continue
        if f.endswith('.csv'):
            df = pd.read_csv(f, dtype=str)
        else:
            df = pd.read_excel(f, dtype=str)
        dfs.append(df.fillna(""))

    if not dfs:
        print("Error: No valid files provided to merge.", file=sys.stderr)
        sys.exit(1)

    combined = pd.concat(dfs, ignore_index=True)
    initial_count = len(combined)

    if dedup_key:
        keys = [k.strip() for k in dedup_key.split(",")]
        valid_keys = [k for k in keys if k in combined.columns]
        if valid_keys:
            combined = combined.drop_duplicates(subset=valid_keys, keep="first")
            print(f"Deduplicated by [{', '.join(valid_keys)}]: {initial_count} → {len(combined)} rows ({initial_count - len(combined)} duplicates removed)")
        else:
            print(f"Warning: None of the dedup keys {keys} were found in columns {list(combined.columns)}", file=sys.stderr)

    os.makedirs(os.path.dirname(os.path.abspath(output_file)), exist_ok=True)
    if output_file.endswith('.csv'):
        combined.to_csv(output_file, index=False)
    else:
        combined.to_excel(output_file, index=False)
    print(f"Successfully merged {len(files)} files into '{output_file}' ({len(combined)} total rows)")

def clean_excel(file_path, trim=True, title_case_cols=None, remove_suffix_cols=None, output_file=None):
    """Clean text fields in Excel file (strip whitespace, titlecase, remove brand suffix)."""
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.", file=sys.stderr)
        sys.exit(1)

    df = pd.read_excel(file_path, dtype=str) if file_path.endswith(('.xlsx', '.xls')) else pd.read_csv(file_path, dtype=str)
    df = df.fillna("")

    if trim:
        for c in df.columns:
            df[c] = df[c].astype(str).str.strip()

    if title_case_cols:
        cols = [c.strip() for c in title_case_cols.split(",")]
        for c in cols:
            if c in df.columns:
                df[c] = df[c].astype(str).str.title()

    out = output_file or file_path
    if out.endswith('.csv'):
        df.to_csv(out, index=False)
    else:
        df.to_excel(out, index=False)
    print(f"Successfully cleaned '{file_path}' → '{out}'")

def main():
    parser = argparse.ArgumentParser(description="Excel ETL Operations Engine (excel-ops.py)")
    subparsers = parser.add_subparsers(dest="command", help="Operation mode")

    read_p = subparsers.add_parser("read", help="Read Excel/CSV data")
    read_p.add_argument("file", help="Path to file")
    read_p.add_argument("--sheet", default=0, help="Sheet name or index")
    read_p.add_argument("--rows", help="Row slice e.g. 10 or 1-50")
    read_p.add_argument("--format", choices=["json", "csv"], default="json", help="Output format")
    read_p.add_argument("--stats", action="store_true", help="Print summary column statistics")

    write_p = subparsers.add_parser("write", help="Write JSON to styled Excel")
    write_p.add_argument("--input", required=True, help="Input JSON file or string")
    write_p.add_argument("--output", required=True, help="Output Excel file path")
    write_p.add_argument("--sheet", default="Data", help="Sheet name")

    merge_p = subparsers.add_parser("merge", help="Merge multiple Excel/CSV files")
    merge_p.add_argument("files", nargs="+", help="Files to merge")
    merge_p.add_argument("--dedup-key", help="Comma-separated column names to deduplicate by")
    merge_p.add_argument("--output", default="merged_output.xlsx", help="Output file path")

    clean_p = subparsers.add_parser("clean", help="Clean string columns in Excel")
    clean_p.add_argument("file", help="File to clean")
    clean_p.add_argument("--trim", action="store_true", default=True, help="Trim whitespace")
    clean_p.add_argument("--title-case", help="Comma-separated columns to titlecase")
    clean_p.add_argument("--output", help="Output file path (overwrites input if not specified)")

    args = parser.parse_args()

    if args.command == "read":
        read_excel(args.file, args.sheet, args.rows, args.format, args.stats)
    elif args.command == "write":
        write_excel(args.input, args.output, args.sheet)
    elif args.command == "merge":
        merge_excel(args.files, args.dedup_key, args.output)
    elif args.command == "clean":
        clean_excel(args.file, args.trim, args.title_case, output_file=args.output)
    else:
        parser.print_help()

if __name__ == "__main__":
    main()
